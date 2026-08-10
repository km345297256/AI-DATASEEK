#!/usr/bin/env python3
"""Create a bounded, deterministic quicklook for common scientific datasets.

The command is deliberately model-free.  It samples tabular and raster inputs,
produces a small set of useful charts, and writes both human- and machine-readable
summaries in one invocation.  ZIP, RAR, and 7z inputs are unpacked through the
same safe recursive extractor shipped with the sandbox.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import asdict, dataclass
from datetime import date, datetime
import hashlib
import io
import json
import math
import os
from pathlib import Path, PurePosixPath
import re
import signal
import shutil
import sys
import tempfile
import threading
import time
from typing import Any, Iterable

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import rasterio
from rasterio.enums import Resampling
import xarray as xr

try:  # Running from sandbox/scripts or from its installed image location.
    from recursive_unpack import Limits as UnpackLimits
    from recursive_unpack import UnpackError, archive_kind, unpack_recursive
except ImportError:  # Importing as scripts.dataset_quicklook in tests.
    from scripts.recursive_unpack import Limits as UnpackLimits
    from scripts.recursive_unpack import UnpackError, archive_kind, unpack_recursive


TABULAR_SUFFIXES = {".csv": "csv", ".tsv": "tsv"}
EXCEL_SUFFIXES = {".xlsx": "excel", ".xlsm": "excel", ".xls": "excel"}
RASTER_SUFFIXES = {".tif": "geotiff", ".tiff": "geotiff"}
NETCDF_SUFFIXES = {".nc": "netcdf", ".nc4": "netcdf", ".cdf": "netcdf"}
SUPPORTED_SUFFIXES = {**TABULAR_SUFFIXES, **EXCEL_SUFFIXES, **RASTER_SUFFIXES, **NETCDF_SUFFIXES}
ARCHIVE_SUFFIXES = {".zip", ".rar", ".7z"}
NULL_TEXT = {"", "na", "n/a", "nan", "null", "none", "-"}
TOOL_EVIDENCE_MAX_DATASETS = 4
TOOL_EVIDENCE_MAX_COLUMNS = 8
TOOL_EVIDENCE_MAX_SHEETS = 2
_TEMPORAL_FIELD_RE = re.compile(
    r"(?:date|time|year|month|day|日期|时间|年份|年度|月份|月)",
    re.IGNORECASE,
)
_LATITUDE_FIELD_RE = re.compile(r"(?:^|[_\s-])(?:lat|latitude|纬度)(?:$|[_\s-])", re.IGNORECASE)
_LONGITUDE_FIELD_RE = re.compile(r"(?:^|[_\s-])(?:lon|lng|longitude|经度)(?:$|[_\s-])", re.IGNORECASE)


class QuicklookError(RuntimeError):
    """Raised when no trustworthy quicklook can be produced."""


@dataclass(frozen=True, slots=True)
class Limits:
    """Hard sampling and discovery limits for a single quicklook invocation."""

    max_files_scan: int = 5_000
    max_files_analyze: int = 24
    max_archives: int = 100
    max_source_bytes: int = 4 * 1024 * 1024 * 1024
    max_text_bytes: int = 8 * 1024 * 1024
    max_rows_per_table: int = 20_000
    max_columns: int = 80
    max_excel_sheets: int = 8
    max_raster_pixels: int = 750_000
    max_raster_bands: int = 8
    max_plot_points: int = 2_000
    max_plots: int = 4
    max_runtime_seconds: int = 90
    max_archive_expanded_bytes: int = 2 * 1024 * 1024 * 1024
    # File-organization output is intentionally smaller than discovery.  It is
    # returned to the conversation, so it must remain readable even when a
    # mounted directory contains thousands of files.
    max_tree_entries: int = 400
    max_tree_depth: int = 12


@dataclass(frozen=True, slots=True)
class Candidate:
    path: Path
    display_path: str
    kind: str
    size: int


@dataclass(slots=True)
class DiscoveryState:
    maximum: int
    files_scanned: int = 0
    supported_files_found: int = 0
    archives_found: int = 0
    truncated: bool = False

    def consume_file(self) -> bool:
        if self.files_scanned >= self.maximum:
            self.truncated = True
            return False
        self.files_scanned += 1
        return True

    def manifest(self, selected: int) -> dict[str, Any]:
        return {
            "files_scanned": self.files_scanned,
            "supported_files_found": self.supported_files_found,
            "archives_found": self.archives_found,
            "files_selected": selected,
            "truncated": self.truncated
            or self.supported_files_found > selected,
        }


@dataclass(slots=True)
class ArchiveBudget:
    archives: int = 0
    extracted_files: int = 0
    expanded_bytes: int = 0


class Deadline:
    def __init__(self, seconds: int) -> None:
        self.started = time.monotonic()
        self.seconds = seconds
        self._old_handler: Any = None
        self._old_timer: tuple[float, float] | None = None
        self._armed = False

    def check(self) -> None:
        if time.monotonic() - self.started > self.seconds:
            raise QuicklookError(
                f"quicklook exceeded its {self.seconds}-second runtime budget"
            )

    @property
    def elapsed(self) -> float:
        return time.monotonic() - self.started

    @property
    def remaining(self) -> float:
        return max(0.0, self.seconds - self.elapsed)

    def arm(self) -> None:
        """Enforce the deadline even while Python is reading a large archive."""
        if (
            self._armed
            or not hasattr(signal, "setitimer")
            or threading.current_thread() is not threading.main_thread()
        ):
            return

        def timeout_handler(_signum: int, _frame: Any) -> None:
            raise QuicklookError(
                f"quicklook exceeded its {self.seconds}-second runtime budget"
            )

        self._old_handler = signal.getsignal(signal.SIGALRM)
        self._old_timer = signal.getitimer(signal.ITIMER_REAL)
        signal.signal(signal.SIGALRM, timeout_handler)
        signal.setitimer(signal.ITIMER_REAL, max(0.001, self.remaining))
        self._armed = True

    def cancel(self) -> None:
        if not self._armed:
            return
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, self._old_handler)
        if self._old_timer and self._old_timer[0] > 0:
            restored = max(0.001, self._old_timer[0] - self.elapsed)
            signal.setitimer(signal.ITIMER_REAL, restored, self._old_timer[1])
        self._armed = False


class PlotCollector:
    def __init__(self, root: Path, maximum: int) -> None:
        self.root = root
        self.maximum = maximum
        self.artifacts: list[dict[str, Any]] = []

    @property
    def remaining(self) -> int:
        return max(0, self.maximum - len(self.artifacts))

    def save(self, figure: Any, *, stem: str, title: str, source: str) -> bool:
        if not self.remaining:
            plt.close(figure)
            return False
        number = len(self.artifacts) + 1
        digest = hashlib.sha1(source.encode("utf-8")).hexdigest()[:8]
        safe_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", stem).strip("_") or "dataset"
        filename = f"quicklook_{number:02d}_{safe_stem[:40]}_{digest}.png"
        temporary = self.root / f".{filename}.tmp"
        final = self.root / filename
        try:
            figure.savefig(
                temporary,
                format="png",
                dpi=140,
                bbox_inches="tight",
                metadata={"Software": "AI-DataSeek dataset quicklook"},
            )
            os.replace(temporary, final)
        finally:
            temporary.unlink(missing_ok=True)
            plt.close(figure)
        self.artifacts.append(
            {
                "path": filename,
                "media_type": "image/png",
                "role": "visualization",
                "title": title,
                "source": source,
                "size": final.stat().st_size,
            }
        )
        return True


def _finite_number(value: Any) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, (np.integer, int)) and not isinstance(value, bool):
        return int(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _json_value(value: Any) -> Any:
    if value is None or value is pd.NA:
        return None
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    if isinstance(value, (np.integer, int)):
        return int(value)
    if isinstance(value, (np.floating, float)):
        return _finite_number(value)
    return str(value)


def _unique_names(values: Iterable[Any], maximum: int) -> list[str]:
    result: list[str] = []
    counts: dict[str, int] = {}
    for index, raw_value in enumerate(values):
        if index >= maximum:
            break
        base = str(raw_value).strip() if raw_value is not None else ""
        base = base or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def _looks_temporal(name: str, values: pd.Series) -> bool:
    if re.search(r"date|time|year|month|day|日期|时间|年份|年度|月份", name, re.I):
        return True
    examples = values.dropna().astype(str).head(20)
    return bool(len(examples)) and examples.str.match(
        r"^\d{4}([-/.]\d{1,2}([-/.]\d{1,2})?)?$"
    ).mean() >= 0.8


def _normalize_frame(frame: pd.DataFrame, limits: Limits) -> pd.DataFrame:
    frame = frame.iloc[: limits.max_rows_per_table, : limits.max_columns].copy()
    frame.columns = _unique_names(frame.columns, limits.max_columns)
    for name in frame.columns:
        series = frame[name]
        if pd.api.types.is_numeric_dtype(series) or pd.api.types.is_datetime64_any_dtype(series):
            continue
        cleaned = series.map(
            lambda value: pd.NA
            if value is None or str(value).strip().lower() in NULL_TEXT
            else str(value).strip()
        )
        present = cleaned.notna()
        if not present.any():
            frame[name] = cleaned
            continue
        numeric = pd.to_numeric(cleaned, errors="coerce")
        if float(numeric.notna().sum()) / int(present.sum()) >= 0.9:
            frame[name] = numeric
            continue
        if _looks_temporal(name, cleaned):
            temporal = pd.to_datetime(cleaned, errors="coerce")
            if float(temporal.notna().sum()) / int(present.sum()) >= 0.8:
                frame[name] = temporal
                continue
        frame[name] = cleaned.astype("string")
    return frame


def _profile_frame(
    frame: pd.DataFrame,
    *,
    rows_total: int | None,
    truncated: bool,
) -> dict[str, Any]:
    row_count = int(len(frame))
    columns: list[dict[str, Any]] = []
    for name in frame.columns:
        series = frame[name]
        missing = int(series.isna().sum())
        column: dict[str, Any] = {
            "name": str(name),
            "dtype": "text",
            "non_null": row_count - missing,
            "missing": missing,
            "missing_percent": round((missing / row_count * 100), 2) if row_count else 0.0,
            "unique_sampled": int(series.nunique(dropna=True)),
        }
        if pd.api.types.is_numeric_dtype(series):
            column["dtype"] = "number"
            numeric = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
            if numeric.notna().any():
                column["statistics"] = {
                    key: _finite_number(value)
                    for key, value in {
                        "min": numeric.min(),
                        "max": numeric.max(),
                        "mean": numeric.mean(),
                        "median": numeric.median(),
                        "std": numeric.std(),
                    }.items()
                }
        elif pd.api.types.is_datetime64_any_dtype(series):
            column["dtype"] = "datetime"
            present = series.dropna()
            if not present.empty:
                column["range"] = {
                    "min": _json_value(present.min()),
                    "max": _json_value(present.max()),
                }
        else:
            counts = series.dropna().astype(str).value_counts().head(5)
            column["top_values"] = [
                {"value": str(value)[:160], "count": int(count)}
                for value, count in counts.items()
            ]
        columns.append(column)
    return {
        "rows_sampled": row_count,
        "rows_total": rows_total,
        "columns_profiled": int(len(frame.columns)),
        "truncated": bool(truncated),
        "columns": columns,
    }


def _decode_text(data: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            pass
    return data.decode("latin-1", errors="replace"), "latin-1"


def _read_delimited(path: Path, kind: str, limits: Limits) -> tuple[pd.DataFrame, dict]:
    with path.open("rb") as stream:
        raw = stream.read(limits.max_text_bytes + 1)
    byte_truncated = len(raw) > limits.max_text_bytes
    raw = raw[: limits.max_text_bytes]
    if byte_truncated:
        boundary = max(raw.rfind(b"\n"), raw.rfind(b"\r"))
        if boundary > 0:
            raw = raw[: boundary + 1]
    text, encoding = _decode_text(raw)
    if not text.strip():
        raise QuicklookError("delimited file is empty")

    delimiter = "\t" if kind == "tsv" else ","
    if kind == "csv":
        try:
            delimiter = csv.Sniffer().sniff(text[:16_384], delimiters=",\t;|").delimiter
        except csv.Error:
            pass
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    try:
        raw_header = next(reader)
    except StopIteration as exc:
        raise QuicklookError("delimited file has no header") from exc
    header = _unique_names(raw_header, limits.max_columns)
    if not header:
        raise QuicklookError("delimited file has no columns")

    rows: list[list[Any]] = []
    row_truncated = False
    for row in reader:
        if len(rows) >= limits.max_rows_per_table:
            row_truncated = True
            break
        selected = row[: len(header)]
        rows.append(selected + [None] * (len(header) - len(selected)))
    frame = _normalize_frame(pd.DataFrame(rows, columns=header), limits)
    return frame, {
        "encoding": encoding,
        "delimiter": delimiter,
        "bytes_read": len(raw),
        "bytes_total": path.stat().st_size,
        "byte_truncated": byte_truncated,
        "row_truncated": row_truncated,
    }


def _rows_to_frame(rows: Iterable[Iterable[Any]], limits: Limits) -> tuple[pd.DataFrame, bool]:
    iterator = iter(rows)
    try:
        header_row = next(iterator)
    except StopIteration:
        return pd.DataFrame(), False
    header = _unique_names(header_row, limits.max_columns)
    if not header:
        return pd.DataFrame(), False
    values: list[list[Any]] = []
    truncated = False
    for row in iterator:
        if len(values) >= limits.max_rows_per_table:
            truncated = True
            break
        selected = list(row)[: len(header)]
        values.append(selected + [None] * (len(header) - len(selected)))
    return _normalize_frame(pd.DataFrame(values, columns=header), limits), truncated


def _read_excel(path: Path, limits: Limits) -> tuple[list[tuple[str, pd.DataFrame]], dict]:
    suffix = path.suffix.lower()
    sheets: list[tuple[str, pd.DataFrame]] = []
    sheet_metadata: list[dict[str, Any]] = []
    if suffix in {".xlsx", ".xlsm"}:
        import openpyxl

        workbook = openpyxl.load_workbook(
            path,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        total_sheets = len(workbook.sheetnames)
        try:
            names = workbook.sheetnames[: limits.max_excel_sheets]
            for name in names:
                worksheet = workbook[name]
                frame, row_truncated = _rows_to_frame(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=limits.max_rows_per_table + 2,
                        max_col=limits.max_columns,
                        values_only=True,
                    ),
                    limits,
                )
                if not frame.columns.empty:
                    sheets.append((name, frame))
                    sheet_metadata.append(
                        {
                            "name": name,
                            "declared_rows": int(worksheet.max_row or 0),
                            "declared_columns": int(worksheet.max_column or 0),
                            "truncated": bool(
                                row_truncated
                                or (worksheet.max_row or 0) > limits.max_rows_per_table + 1
                                or (worksheet.max_column or 0) > limits.max_columns
                            ),
                        }
                    )
        finally:
            workbook.close()
    else:
        excel_file = pd.ExcelFile(path, engine="xlrd")
        total_sheets = len(excel_file.sheet_names)
        for name in excel_file.sheet_names[: limits.max_excel_sheets]:
            frame = excel_file.parse(name, nrows=limits.max_rows_per_table + 1)
            truncated = len(frame) > limits.max_rows_per_table or len(frame.columns) > limits.max_columns
            frame = _normalize_frame(frame, limits)
            if not frame.columns.empty:
                sheets.append((name, frame))
                sheet_metadata.append({"name": name, "truncated": truncated})
    if not sheets:
        raise QuicklookError("workbook contains no readable tabular sheets")
    return sheets, {
        "sheet_count": total_sheets,
        "sheets_profiled": len(sheets),
        "sheets_truncated": total_sheets > limits.max_excel_sheets,
        "sheet_metadata": sheet_metadata,
    }


def _sample_frame(frame: pd.DataFrame, maximum: int) -> pd.DataFrame:
    if len(frame) <= maximum:
        return frame
    indexes = np.linspace(0, len(frame) - 1, maximum, dtype=int)
    return frame.iloc[indexes]


def _plot_table_primary(
    frame: pd.DataFrame,
    *,
    label: str,
    stem: str,
    source: str,
    plots: PlotCollector,
    limits: Limits,
) -> None:
    if not plots.remaining:
        return
    numeric_names = list(frame.select_dtypes(include=[np.number]).columns)[:4]
    if numeric_names:
        rows = math.ceil(len(numeric_names) / 2)
        figure, axes = plt.subplots(rows, 2, figsize=(10, 3.2 * rows), squeeze=False)
        for axis, name in zip(axes.flat, numeric_names):
            values = pd.to_numeric(frame[name], errors="coerce").replace(
                [np.inf, -np.inf], np.nan
            ).dropna()
            values = _sample_frame(values.to_frame(), limits.max_plot_points).iloc[:, 0]
            if values.empty:
                axis.text(0.5, 0.5, "无有效数值", ha="center", va="center")
            else:
                axis.hist(values, bins=min(30, max(5, int(math.sqrt(len(values))))), color="#267a63", alpha=0.85)
            axis.set_title(str(name))
            axis.set_ylabel("频数")
            axis.grid(alpha=0.2)
        for axis in list(axes.flat)[len(numeric_names):]:
            axis.set_visible(False)
        figure.suptitle(f"{label} · 数值分布")
        figure.tight_layout()
        plots.save(
            figure,
            stem=f"{stem}_distribution",
            title=f"{label} · 数值分布",
            source=source,
        )
        return

    text_names = list(frame.columns)[:4]
    figure, axes = plt.subplots(max(1, len(text_names)), 1, figsize=(10, 3 * max(1, len(text_names))), squeeze=False)
    for axis, name in zip(axes.flat, text_names):
        counts = frame[name].dropna().astype(str).value_counts().head(10).sort_values()
        if counts.empty:
            axis.text(0.5, 0.5, "无有效记录", ha="center", va="center")
            axis.set_axis_off()
        else:
            labels = [value[:28] + ("…" if len(value) > 28 else "") for value in counts.index]
            axis.barh(labels, counts.values, color="#267a63", alpha=0.85)
            axis.set_title(str(name))
            axis.set_xlabel("频数")
            axis.grid(axis="x", alpha=0.2)
    figure.suptitle(f"{label} · 类别概览")
    figure.tight_layout()
    plots.save(
        figure,
        stem=f"{stem}_categories",
        title=f"{label} · 类别概览",
        source=source,
    )


def _plot_table_extras(
    frame: pd.DataFrame,
    *,
    label: str,
    stem: str,
    source: str,
    plots: PlotCollector,
    limits: Limits,
) -> None:
    if not plots.remaining:
        return
    numeric = frame.select_dtypes(include=[np.number]).iloc[:, :12]
    if numeric.shape[1] >= 2:
        correlation = _sample_frame(numeric, limits.max_plot_points).corr(min_periods=3)
        figure, axis = plt.subplots(figsize=(max(6, correlation.shape[1] * 0.65), max(5, correlation.shape[0] * 0.55)))
        image = axis.imshow(correlation, vmin=-1, vmax=1, cmap="RdBu_r")
        axis.set_xticks(range(len(correlation.columns)), correlation.columns, rotation=45, ha="right")
        axis.set_yticks(range(len(correlation.columns)), correlation.columns)
        axis.set_title(f"{label} · 数值相关性")
        figure.colorbar(image, ax=axis, label="相关系数")
        figure.tight_layout()
        plots.save(figure, stem=f"{stem}_correlation", title=f"{label} · 数值相关性", source=source)

    if plots.remaining:
        missing = frame.isna().mean().sort_values(ascending=False)
        missing = missing[missing > 0].head(20).sort_values()
        if not missing.empty:
            figure, axis = plt.subplots(figsize=(9, max(4, len(missing) * 0.32)))
            axis.barh(missing.index, missing.values * 100, color="#e0a23b")
            axis.set_xlabel("缺失率（%）")
            axis.set_title(f"{label} · 缺失值概览")
            axis.grid(axis="x", alpha=0.2)
            figure.tight_layout()
            plots.save(figure, stem=f"{stem}_missing", title=f"{label} · 缺失值概览", source=source)

    if plots.remaining and not numeric.empty:
        temporal_names = [
            name for name in frame.columns if pd.api.types.is_datetime64_any_dtype(frame[name])
        ]
        if temporal_names:
            time_name = temporal_names[0]
            value_name = numeric.columns[0]
            selected = frame[[time_name, value_name]].dropna().sort_values(time_name)
            selected = _sample_frame(selected, limits.max_plot_points)
            if not selected.empty:
                figure, axis = plt.subplots(figsize=(10, 4.5))
                axis.plot(selected[time_name], selected[value_name], color="#267a63", linewidth=1.6)
                axis.set_xlabel(str(time_name))
                axis.set_ylabel(str(value_name))
                axis.set_title(f"{label} · 时间趋势")
                axis.grid(alpha=0.2)
                figure.autofmt_xdate()
                figure.tight_layout()
                plots.save(figure, stem=f"{stem}_trend", title=f"{label} · 时间趋势", source=source)


def _netcdf_profile_and_plot(
    candidate: Candidate,
    plots: PlotCollector,
    limits: Limits,
) -> dict[str, Any]:
    """Profile one NetCDF file with bounded reads and one representative chart."""
    try:
        dataset = xr.open_dataset(candidate.path, engine="h5netcdf", chunks={})
    except Exception as exc:
        raise QuicklookError(f"无法读取 NetCDF 文件: {type(exc).__name__}: {exc}") from exc
    try:
        variables: list[dict[str, Any]] = []
        numeric_variables: list[str] = []
        for name, variable in dataset.variables.items():
            variables.append(
                {
                    "name": str(name),
                    "dimensions": list(variable.dims),
                    "shape": [int(size) for size in variable.shape],
                    "dtype": str(variable.dtype),
                    "units": str(variable.attrs.get("units", ""))[:120],
                    "long_name": str(variable.attrs.get("long_name", ""))[:200],
                }
            )
            if name not in dataset.coords and np.issubdtype(variable.dtype, np.number):
                numeric_variables.append(name)
        if not numeric_variables:
            raise QuicklookError("NetCDF 文件中没有可分析的数值变量")

        selected = numeric_variables[0]
        data = dataset[selected]
        # Bound the product of sampled dimensions, not just each dimension
        # independently (a 3-D climate cube would otherwise explode).
        dimensions = list(data.dims)
        target = max(1, min(limits.max_raster_pixels, limits.max_plot_points * 4))
        per_dimension = max(1, int(target ** (1 / max(1, len(dimensions)))))
        for dimension in dimensions:
            size = int(data.sizes[dimension])
            sample_size = min(size, per_dimension)
            if sample_size < size:
                indexes = np.linspace(0, size - 1, sample_size, dtype=int)
                data = data.isel({dimension: indexes})
        values = np.asarray(data.values, dtype=float)
        finite = values[np.isfinite(values)]
        if plots.remaining and finite.size:
            figure, axis = plt.subplots(figsize=(9, 4.8))
            sampled = finite
            if sampled.size > limits.max_plot_points:
                sampled = np.asarray(
                    _sample_frame(pd.DataFrame({"value": sampled}), limits.max_plot_points)["value"]
                )
            axis.hist(
                sampled,
                bins=min(40, max(8, int(math.sqrt(sampled.size)))),
                color="#267a63",
                alpha=0.85,
            )
            axis.set_title(f"{candidate.path.stem}: {selected}")
            axis.set_xlabel(str(dataset[selected].attrs.get("units", "value")))
            axis.set_ylabel("频数")
            axis.grid(alpha=0.2)
            figure.tight_layout()
            plots.save(
                figure,
                stem=f"{candidate.path.stem}_{selected}_distribution",
                title=f"{candidate.path.stem}: {selected}",
                source=candidate.display_path,
            )
        return {
            "path": candidate.display_path,
            "format": "netcdf",
            "size": candidate.size,
            "dimensions": {str(name): int(size) for name, size in dataset.sizes.items()},
            "variables": variables[: limits.max_columns],
            "variables_truncated": len(variables) > limits.max_columns,
            "sampled_variable": selected,
            "sampled_values": int(finite.size),
        }
    finally:
        dataset.close()


def _scaled_shape(height: int, width: int, maximum_pixels: int) -> tuple[int, int]:
    pixels = max(1, height * width)
    scale = max(1.0, math.sqrt(pixels / maximum_pixels))
    return max(1, int(height / scale)), max(1, int(width / scale))


def _declared_nodata_value(value: Any) -> Any:
    """Render a declared NoData value without conflating NaN with no declaration."""
    if value is None:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return _json_value(value)
    if math.isnan(number):
        return "NaN"
    if math.isinf(number):
        return "Infinity" if number > 0 else "-Infinity"
    return _json_value(value)


def _raster_mask_provenance(dataset: Any, band_number: int) -> list[str]:
    """Return Rasterio/GDAL mask sources for one band in a stable JSON form."""
    try:
        flags = dataset.mask_flag_enums[band_number - 1]
    except (AttributeError, IndexError, TypeError):
        return ["unknown"]

    names: list[str] = []
    for flag in flags:
        name = getattr(flag, "name", None)
        if not name:
            name = str(flag).rsplit(".", 1)[-1]
        normalized = str(name).strip().lower()
        if normalized and normalized not in names:
            names.append(normalized)
    order = {"nodata": 0, "alpha": 1, "per_dataset": 2, "all_valid": 3}
    return sorted(names or ["unknown"], key=lambda item: (order.get(item, 99), item))


def _raster_declared_unit(dataset: Any, band_number: int) -> str | None:
    """Read only the unit explicitly declared in Rasterio/GDAL band metadata."""
    try:
        value = dataset.units[band_number - 1]
    except (AttributeError, IndexError, TypeError):
        return None
    if value is None:
        return None
    declared = str(value).strip()
    return declared[:160] or None


def _raster_spatial_profile(
    data: np.ma.MaskedArray,
    dataset: Any,
) -> dict[str, Any]:
    """Summarize sampled spatial structure without inventing geography.

    Zone labels describe the raster grid (upper/lower, left/right). They are not
    called north/south unless a later analyst verifies the affine orientation.
    """
    values = np.asarray(data.astype(np.float64).filled(np.nan), dtype=np.float64)
    finite_mask = np.isfinite(values)
    finite_values = values[finite_mask]
    if not finite_values.size:
        return {
            "sampled_grid": {"height": int(values.shape[0]), "width": int(values.shape[1])},
            "valid_pixels": 0,
            "valid_fraction_percent": 0.0,
            "quantiles": {},
            "zone_means": {},
            "row_band_means": {},
            "column_band_means": {},
            "minimum_location": None,
            "maximum_location": None,
        }

    def mean_of(region: np.ndarray) -> float | None:
        finite = region[np.isfinite(region)]
        return _finite_number(finite.mean()) if finite.size else None

    height, width = values.shape
    row_mid = max(1, height // 2)
    col_mid = max(1, width // 2)
    row_third = max(1, height // 3)
    col_third = max(1, width // 3)

    def original_location(sample_row: int, sample_col: int, value: float) -> dict[str, Any]:
        original_row = min(
            dataset.height - 1,
            max(0, int(round((sample_row + 0.5) * dataset.height / height - 0.5))),
        )
        original_col = min(
            dataset.width - 1,
            max(0, int(round((sample_col + 0.5) * dataset.width / width - 0.5))),
        )
        x, y = dataset.xy(original_row, original_col)
        return {
            "value": _finite_number(value),
            "row": original_row,
            "column": original_col,
            "x": _finite_number(x),
            "y": _finite_number(y),
        }

    finite_grid = np.where(finite_mask, values, np.nan)
    minimum_index = np.unravel_index(np.nanargmin(finite_grid), values.shape)
    maximum_index = np.unravel_index(np.nanargmax(finite_grid), values.shape)
    quantile_values = np.quantile(finite_values, [0.05, 0.25, 0.5, 0.75, 0.95])
    return {
        "sampled_grid": {"height": int(height), "width": int(width)},
        "valid_pixels": int(finite_values.size),
        "valid_fraction_percent": round(
            float(finite_values.size / max(1, values.size) * 100),
            3,
        ),
        "quantiles": {
            key: _finite_number(value)
            for key, value in zip(
                ("p05", "p25", "p50", "p75", "p95"),
                quantile_values,
            )
        },
        "zone_means": {
            "upper_left": mean_of(values[:row_mid, :col_mid]),
            "upper_right": mean_of(values[:row_mid, col_mid:]),
            "lower_left": mean_of(values[row_mid:, :col_mid]),
            "lower_right": mean_of(values[row_mid:, col_mid:]),
            "center": mean_of(
                values[
                    row_third : max(row_third + 1, height - row_third),
                    col_third : max(col_third + 1, width - col_third),
                ]
            ),
        },
        "row_band_means": {
            "upper_third": mean_of(values[:row_third, :]),
            "middle_third": mean_of(
                values[row_third : max(row_third + 1, height - row_third), :]
            ),
            "lower_third": mean_of(values[max(row_third, height - row_third) :, :]),
        },
        "column_band_means": {
            "left_third": mean_of(values[:, :col_third]),
            "middle_third": mean_of(
                values[:, col_third : max(col_third + 1, width - col_third)]
            ),
            "right_third": mean_of(values[:, max(col_third, width - col_third) :]),
        },
        "minimum_location": original_location(
            int(minimum_index[0]),
            int(minimum_index[1]),
            float(values[minimum_index]),
        ),
        "maximum_location": original_location(
            int(maximum_index[0]),
            int(maximum_index[1]),
            float(values[maximum_index]),
        ),
    }


def _raster_profile_and_plot(
    candidate: Candidate,
    plots: PlotCollector,
    limits: Limits,
) -> dict[str, Any]:
    with rasterio.open(candidate.path) as dataset:
        sampled_height, sampled_width = _scaled_shape(
            dataset.height, dataset.width, limits.max_raster_pixels
        )
        band_profiles: list[dict[str, Any]] = []
        sampled_bands = min(dataset.count, limits.max_raster_bands)
        first_band: np.ma.MaskedArray | None = None
        for band_number in range(1, sampled_bands + 1):
            data = dataset.read(
                band_number,
                out_shape=(sampled_height, sampled_width),
                masked=True,
                resampling=Resampling.nearest,
            )
            if first_band is None:
                first_band = data
            raw_values = np.asarray(data.data)
            authoritative_mask = np.ma.getmaskarray(data)
            raw_zero_mask = np.equal(raw_values, 0)
            values = np.asarray(data.compressed(), dtype=np.float64)
            values = values[np.isfinite(values)]
            band_profiles.append(
                {
                    "band": band_number,
                    "description": dataset.descriptions[band_number - 1],
                    "declared_nodata": _declared_nodata_value(
                        dataset.nodatavals[band_number - 1]
                    ),
                    "declared_unit": _raster_declared_unit(dataset, band_number),
                    "mask_provenance": _raster_mask_provenance(
                        dataset, band_number
                    ),
                    "masked_count": int(np.count_nonzero(authoritative_mask)),
                    "nan_count": int(np.count_nonzero(np.isnan(raw_values))),
                    # zero_count describes what is physically present in the
                    # sampled cells. valid_zero_count makes explicit how many
                    # of those zeros Rasterio's authoritative mask retained.
                    "zero_count": int(np.count_nonzero(raw_zero_mask)),
                    "valid_zero_count": int(
                        np.count_nonzero(raw_zero_mask & ~authoritative_mask)
                    ),
                    "valid_pixels_sampled": int(values.size),
                    "min": _finite_number(values.min()) if values.size else None,
                    "max": _finite_number(values.max()) if values.size else None,
                    "mean": _finite_number(values.mean()) if values.size else None,
                    "std": _finite_number(values.std()) if values.size else None,
                }
            )

        if plots.remaining and first_band is not None:
            figure, axis = plt.subplots(figsize=(9, 6))
            image = axis.imshow(first_band, cmap="viridis")
            axis.set_title(f"{Path(candidate.display_path).name} · 栅格快视图")
            axis.set_xlabel("列")
            axis.set_ylabel("行")
            figure.colorbar(image, ax=axis, label=dataset.descriptions[0] or "像元值")
            figure.tight_layout()
            plots.save(
                figure,
                stem=f"{candidate.path.stem}_raster",
                title=f"{Path(candidate.display_path).name} · 栅格快视图",
                source=candidate.display_path,
            )

        if plots.remaining and first_band is not None:
            values = np.asarray(first_band.compressed(), dtype=np.float64)
            values = values[np.isfinite(values)]
            if values.size:
                if values.size > limits.max_plot_points:
                    indexes = np.linspace(
                        0, values.size - 1, limits.max_plot_points, dtype=int
                    )
                    values = values[indexes]
                sorted_values = np.sort(values)
                cumulative = np.arange(1, len(sorted_values) + 1) / len(sorted_values)
                figure, axes = plt.subplots(1, 2, figsize=(11, 4.5))
                axes[0].hist(
                    values,
                    bins=min(40, max(8, int(math.sqrt(len(values))))),
                    color="#267a63",
                    alpha=0.85,
                )
                axes[0].set_title("像元值分布")
                axes[0].set_xlabel("像元值")
                axes[0].set_ylabel("频数")
                axes[0].grid(alpha=0.2)
                axes[1].plot(sorted_values, cumulative, color="#d0782d", linewidth=1.8)
                axes[1].set_title("累积分布（CDF）")
                axes[1].set_xlabel("像元值")
                axes[1].set_ylabel("累计比例")
                axes[1].grid(alpha=0.2)
                title = f"{Path(candidate.display_path).name} · 数值分布与 CDF"
                figure.suptitle(title)
                figure.tight_layout()
                plots.save(
                    figure,
                    stem=f"{candidate.path.stem}_distribution_cdf",
                    title=title,
                    source=candidate.display_path,
                )

        if plots.remaining and first_band is not None and first_band.count():
            row_profile = np.ma.mean(first_band, axis=1).filled(np.nan)
            column_profile = np.ma.mean(first_band, axis=0).filled(np.nan)
            row_positions = np.linspace(0, dataset.height - 1, len(row_profile))
            column_positions = np.linspace(0, dataset.width - 1, len(column_profile))
            if len(row_profile) > limits.max_plot_points:
                indexes = np.linspace(
                    0, len(row_profile) - 1, limits.max_plot_points, dtype=int
                )
                row_profile = row_profile[indexes]
                row_positions = row_positions[indexes]
            if len(column_profile) > limits.max_plot_points:
                indexes = np.linspace(
                    0, len(column_profile) - 1, limits.max_plot_points, dtype=int
                )
                column_profile = column_profile[indexes]
                column_positions = column_positions[indexes]
            figure, axes = plt.subplots(1, 2, figsize=(11, 4.5))
            axes[0].plot(row_positions, row_profile, color="#267a63", linewidth=1.5)
            axes[0].set_title("逐行均值剖面")
            axes[0].set_xlabel("原始行号")
            axes[0].set_ylabel("平均像元值")
            axes[0].grid(alpha=0.2)
            axes[1].plot(
                column_positions,
                column_profile,
                color="#d0782d",
                linewidth=1.5,
            )
            axes[1].set_title("逐列均值剖面")
            axes[1].set_xlabel("原始列号")
            axes[1].set_ylabel("平均像元值")
            axes[1].grid(alpha=0.2)
            title = f"{Path(candidate.display_path).name} · 行列剖面"
            figure.suptitle(title)
            figure.tight_layout()
            plots.save(
                figure,
                stem=f"{candidate.path.stem}_profiles",
                title=title,
                source=candidate.display_path,
            )

        spatial_profile = (
            _raster_spatial_profile(first_band, dataset)
            if first_band is not None
            else None
        )
        declared_nodata_by_band = [
            _declared_nodata_value(dataset.nodatavals[index])
            for index in range(sampled_bands)
        ]
        declared_nodata: Any = (
            declared_nodata_by_band[0]
            if declared_nodata_by_band
            and all(value == declared_nodata_by_band[0] for value in declared_nodata_by_band)
            else declared_nodata_by_band
        )
        declared_units_by_band = [band["declared_unit"] for band in band_profiles]
        declared_unit: Any = (
            declared_units_by_band[0]
            if declared_units_by_band
            and all(value == declared_units_by_band[0] for value in declared_units_by_band)
            else declared_units_by_band
        )
        mask_provenance = sorted(
            {
                provenance
                for band in band_profiles
                for provenance in band["mask_provenance"]
            },
            key=lambda item: (
                {"nodata": 0, "alpha": 1, "per_dataset": 2, "all_valid": 3}.get(
                    item, 99
                ),
                item,
            ),
        )
        return {
            "path": candidate.display_path,
            "format": "geotiff",
            "size": candidate.size,
            "width": dataset.width,
            "height": dataset.height,
            "band_count": dataset.count,
            "bands_profiled": sampled_bands,
            "dtypes": list(dataset.dtypes),
            "crs": str(dataset.crs) if dataset.crs else None,
            "bounds": [
                _finite_number(value)
                for value in (dataset.bounds.left, dataset.bounds.bottom, dataset.bounds.right, dataset.bounds.top)
            ],
            "nodata": _json_value(dataset.nodata),
            "declared_nodata": declared_nodata,
            "declared_unit": declared_unit,
            "mask_provenance": mask_provenance,
            "masked_count": sum(band["masked_count"] for band in band_profiles),
            "nan_count": sum(band["nan_count"] for band in band_profiles),
            "zero_count": sum(band["zero_count"] for band in band_profiles),
            "valid_zero_count": sum(
                band["valid_zero_count"] for band in band_profiles
            ),
            "sampling": {
                "width": sampled_width,
                "height": sampled_height,
                "pixels_per_band": sampled_width * sampled_height,
                "truncated": sampled_width != dataset.width
                or sampled_height != dataset.height
                or sampled_bands != dataset.count,
            },
            "bands": band_profiles,
            "spatial_profile": spatial_profile,
        }


def _safe_relative_manifest_path(raw_path: Any) -> str | None:
    """Normalize an extractor path without ever accepting an absolute path."""
    text = str(raw_path).replace("\\", "/").strip()
    path = PurePosixPath(text)
    if not text or path.is_absolute() or ".." in path.parts:
        return None
    parts = [part for part in path.parts if part not in {"", "."}]
    return "/".join(parts) if parts else "."


def _public_error_text(
    exc: BaseException,
    replacements: Iterable[tuple[Path, str]],
    *,
    include_type: bool = True,
) -> str:
    """Redact sandbox paths from errors that may become public artifacts."""
    text = str(exc)
    private_values: list[tuple[str, str]] = []
    for private, public in replacements:
        private_values.append((str(private), public))
        try:
            private_values.append((str(private.resolve(strict=False)), public))
        except OSError:
            pass
    for private, public in sorted(set(private_values), key=lambda item: -len(item[0])):
        if private:
            text = text.replace(private, public)
    rendered = f"{type(exc).__name__}: {text}" if include_type else text
    return rendered[:1_000]


def _source_root_description(source: Path) -> dict[str, Any]:
    name = source.name or "dataset"
    if source.is_dir():
        return {"name": name, "type": "directory"}
    archive_format = archive_kind(source)
    if archive_format is not None:
        return {
            "name": name,
            "type": "archive",
            "format": archive_format,
            "size": source.stat().st_size,
        }
    return {
        "name": name,
        "type": "file",
        "format": _kind_for_path(source) or "unknown",
        "size": source.stat().st_size,
    }


def _original_source_tree(source: Path, limits: Limits) -> dict[str, Any]:
    """Return a bounded, source-relative listing suitable for public output."""
    root = _source_root_description(source)
    if not source.is_dir():
        return {"root": root, "entries": [], "entries_listed": 0, "truncated": False}

    entries: list[dict[str, Any]] = []
    truncated = False
    source = source.resolve()
    for directory, directory_names, filenames in os.walk(source, followlinks=False):
        directory_path = Path(directory)
        relative_directory = directory_path.relative_to(source)
        directory_depth = 0 if relative_directory == Path(".") else len(relative_directory.parts)
        if directory_depth >= limits.max_tree_depth:
            if directory_names:
                truncated = True
            directory_names[:] = []
        else:
            directory_names[:] = sorted(
                name
                for name in directory_names
                if not (directory_path / name).is_symlink()
            )

        for name in directory_names:
            if len(entries) >= limits.max_tree_entries:
                truncated = True
                break
            path = directory_path / name
            relative = path.relative_to(source).as_posix()
            entries.append({"path": relative, "type": "directory"})
        if truncated and len(entries) >= limits.max_tree_entries:
            break

        for name in sorted(filenames):
            if len(entries) >= limits.max_tree_entries:
                truncated = True
                break
            path = directory_path / name
            if path.is_symlink() or not path.is_file():
                continue
            relative = path.relative_to(source).as_posix()
            suffix = path.suffix.lower()
            if suffix in ARCHIVE_SUFFIXES:
                entry_type = "archive"
                file_format = suffix.removeprefix(".")
            else:
                entry_type = "file"
                file_format = _kind_for_path(path)
            entry: dict[str, Any] = {
                "path": relative,
                "type": entry_type,
                "size": path.stat().st_size,
            }
            if file_format:
                entry["format"] = file_format
            entries.append(entry)
        if truncated and len(entries) >= limits.max_tree_entries:
            break

    return {
        "root": root,
        "entries": entries,
        "entries_listed": len(entries),
        "truncated": truncated,
    }


def _archive_organization(
    unpack_manifest: dict[str, Any],
    *,
    outer_display_path: str,
    max_layers: int,
    max_files: int,
) -> dict[str, Any]:
    """Convert a private unpack manifest to bounded logical public paths."""
    archive_layers: list[dict[str, Any]] = []
    extracted_entries: list[dict[str, Any]] = []
    raw_archives = unpack_manifest.get("archives", [])
    raw_files = unpack_manifest.get("files", [])

    for raw in raw_archives[:max_layers]:
        depth = max(0, int(raw.get("depth", 0)))
        internal = _safe_relative_manifest_path(raw.get("path"))
        if internal is None:
            continue
        path = outer_display_path if depth == 0 else f"{outer_display_path}!/{internal}"
        extracted_to = _safe_relative_manifest_path(raw.get("extracted_to", "."))
        if extracted_to is None or extracted_to == ".":
            logical_target = f"{outer_display_path}!/"
        else:
            logical_target = f"{outer_display_path}!/{extracted_to}/"
        archive_layers.append(
            {
                "path": path,
                "format": str(raw.get("format", "archive")),
                "level": depth + 1,
                "extracted_to": logical_target,
            }
        )

    for raw in raw_files[:max_files]:
        internal = _safe_relative_manifest_path(raw.get("path"))
        if internal is None or internal == ".":
            continue
        suffix = PurePosixPath(internal).suffix.lower()
        entry: dict[str, Any] = {
            "path": f"{outer_display_path}!/{internal}",
            "type": "file",
            "size": max(0, int(raw.get("size", 0))),
        }
        file_format = SUPPORTED_SUFFIXES.get(suffix)
        if file_format:
            entry["format"] = file_format
        extracted_entries.append(entry)

    return {
        "archive_layers": archive_layers,
        "archive_layers_truncated": len(raw_archives) > len(archive_layers),
        "extracted_entries": extracted_entries,
        "extracted_tree_truncated": len(raw_files) > len(extracted_entries),
    }


def _kind_for_path(path: Path) -> str | None:
    return SUPPORTED_SUFFIXES.get(path.suffix.lower())


def _discover(
    root: Path,
    *,
    excluded: Path,
    limits: Limits,
    state: DiscoveryState,
    include_archives: bool = False,
    display_prefix: str | None = None,
) -> tuple[list[Candidate], list[Candidate]]:
    candidates: list[Candidate] = []
    archives: list[Candidate] = []
    root = root.resolve()
    excluded = excluded.resolve()
    for directory, directory_names, filenames in os.walk(root, followlinks=False):
        directory_path = Path(directory)
        directory_names[:] = sorted(
            name
            for name in directory_names
            if not (directory_path / name).is_symlink()
            and (directory_path / name).resolve() != excluded
        )
        for name in sorted(filenames):
            if not state.consume_file():
                break
            path = directory_path / name
            if path.is_symlink() or not path.is_file():
                continue
            relative_path = path.relative_to(root).as_posix()
            shown_path = (
                f"{display_prefix}!/{relative_path}"
                if display_prefix
                else relative_path
            )
            if include_archives and path.suffix.lower() in ARCHIVE_SUFFIXES:
                state.archives_found += 1
                archives.append(
                    Candidate(
                        path=path,
                        display_path=shown_path,
                        kind="archive",
                        size=path.stat().st_size,
                    )
                )
                continue
            kind = _kind_for_path(path)
            if kind is None:
                continue
            size = path.stat().st_size
            state.supported_files_found += 1
            candidates.append(
                Candidate(
                    path=path,
                    display_path=shown_path,
                    kind=kind,
                    size=size,
                )
            )
        if state.truncated:
            break
    candidates.sort(key=lambda item: item.display_path.casefold())
    archives.sort(key=lambda item: item.display_path.casefold())
    return candidates, archives


def _archive_output_name(index: int, display_path: str) -> str:
    stem = Path(display_path).stem
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._") or "archive"
    digest = hashlib.sha1(display_path.encode("utf-8")).hexdigest()[:8]
    return f"{index:03d}_{safe_stem[:40]}_{digest}"


def _unpack_directory_archives(
    archives: list[Candidate],
    *,
    staging: Path,
    limits: Limits,
    deadline: Deadline,
    discovery_state: DiscoveryState,
    selected_count: int,
) -> tuple[
    list[Candidate],
    dict[str, Any],
    list[dict[str, str]],
    dict[str, Any],
]:
    """Unpack directory archives under one shared global resource budget."""
    budget = ArchiveBudget()
    selected: list[Candidate] = []
    items: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    archive_layers: list[dict[str, Any]] = []
    extracted_entries: list[dict[str, Any]] = []
    archive_layers_truncated = False
    extracted_tree_truncated = False
    for index, archive in enumerate(archives, start=1):
        deadline.check()
        if selected_count + len(selected) >= limits.max_files_analyze:
            items.append(
                {
                    "path": archive.display_path,
                    "status": "skipped",
                    "reason": "analysis file limit reached",
                }
            )
            continue
        if budget.archives >= limits.max_archives:
            items.append(
                {
                    "path": archive.display_path,
                    "status": "skipped",
                    "reason": "global archive count limit reached",
                }
            )
            continue
        if archive.size > limits.max_source_bytes:
            error = (
                f"archive exceeds size limit ({archive.size} > "
                f"{limits.max_source_bytes} bytes)"
            )
            items.append(
                {"path": archive.display_path, "status": "failed", "error": error}
            )
            errors.append({"path": archive.display_path, "error": error})
            budget.archives += 1
            continue

        remaining_archives = limits.max_archives - budget.archives
        remaining_files = limits.max_files_scan - budget.extracted_files
        remaining_bytes = limits.max_archive_expanded_bytes - budget.expanded_bytes
        if remaining_files <= 0 or remaining_bytes <= 0:
            items.append(
                {
                    "path": archive.display_path,
                    "status": "skipped",
                    "reason": "global archive expansion budget reached",
                }
            )
            continue
        target = staging / "unpacked_archives" / _archive_output_name(
            index, archive.display_path
        )
        try:
            unpack_manifest = unpack_recursive(
                archive.path,
                target,
                UnpackLimits(
                    max_depth=5,
                    max_archives=remaining_archives,
                    max_files=remaining_files,
                    max_total_bytes=remaining_bytes,
                    max_single_file_bytes=min(
                        limits.max_source_bytes,
                        limits.max_archive_expanded_bytes,
                    ),
                    timeout_seconds=max(
                        1, min(int(math.ceil(deadline.remaining)), 600)
                    ),
                ),
            )
        except QuicklookError:
            raise
        except Exception as exc:
            budget.archives += 1
            error = _public_error_text(
                exc,
                (
                    (archive.path, archive.display_path),
                    (target, "[内部解压目录]"),
                    (staging, "[内部工作目录]"),
                ),
            )
            items.append(
                {"path": archive.display_path, "status": "failed", "error": error}
            )
            errors.append({"path": archive.display_path, "error": error})
            continue

        summary = unpack_manifest["summary"]
        archive_count = int(summary["archive_count"])
        final_file_count = int(summary["file_count"])
        extracted_file_count = final_file_count + max(0, archive_count - 1)
        expanded_bytes = int(summary["expanded_bytes"])
        budget.archives += archive_count
        budget.extracted_files += extracted_file_count
        budget.expanded_bytes += expanded_bytes

        organization = _archive_organization(
            unpack_manifest,
            outer_display_path=archive.display_path,
            max_layers=max(0, limits.max_tree_entries - len(archive_layers)),
            max_files=max(0, limits.max_tree_entries - len(extracted_entries)),
        )
        archive_layers.extend(organization["archive_layers"])
        extracted_entries.extend(organization["extracted_entries"])
        archive_layers_truncated = (
            archive_layers_truncated
            or organization["archive_layers_truncated"]
        )
        extracted_tree_truncated = (
            extracted_tree_truncated
            or organization["extracted_tree_truncated"]
        )

        discovered, _ignored_archives = _discover(
            target,
            excluded=staging,
            limits=limits,
            state=discovery_state,
            include_archives=False,
            display_prefix=archive.display_path,
        )
        remaining_selection = max(
            0, limits.max_files_analyze - selected_count - len(selected)
        )
        selected.extend(discovered[:remaining_selection])
        items.append(
            {
                "path": archive.display_path,
                "status": "processed",
                "archive_count": archive_count,
                "final_file_count": final_file_count,
                "extracted_file_count": extracted_file_count,
                "expanded_bytes": expanded_bytes,
                "supported_files_found": len(discovered),
            }
        )

    return (
        selected,
        {
            "discovered": len(archives),
            "processed": sum(item["status"] == "processed" for item in items),
            "failed": sum(item["status"] == "failed" for item in items),
            "skipped": sum(item["status"] == "skipped" for item in items),
            "archive_count": budget.archives,
            "extracted_file_count": budget.extracted_files,
            "expanded_bytes": budget.expanded_bytes,
            "items": items,
        },
        errors,
        {
            "archive_layers": archive_layers,
            "archive_layers_truncated": archive_layers_truncated,
            "extracted_entries": extracted_entries,
            "extracted_tree_truncated": extracted_tree_truncated,
        },
    )


def _single_candidate(path: Path, limits: Limits) -> Candidate:
    kind = _kind_for_path(path)
    if kind is None:
        raise QuicklookError(
            "unsupported input; expected a directory, CSV/TSV, Excel, NetCDF, GeoTIFF, "
            "ZIP, RAR, or 7z"
        )
    size = path.stat().st_size
    if size > limits.max_source_bytes:
        raise QuicklookError(
            f"source exceeds size limit ({size} > {limits.max_source_bytes} bytes)"
        )
    return Candidate(path=path, display_path=path.name, kind=kind, size=size)


def _analyze_table_candidate(
    candidate: Candidate,
    plots: PlotCollector,
    limits: Limits,
) -> tuple[dict[str, Any], tuple[pd.DataFrame, str, str] | None]:
    label = Path(candidate.display_path).name
    if candidate.kind in {"csv", "tsv"}:
        frame, read_metadata = _read_delimited(candidate.path, candidate.kind, limits)
        profile = _profile_frame(
            frame,
            rows_total=None,
            truncated=read_metadata["byte_truncated"] or read_metadata["row_truncated"],
        )
        result = {
            "path": candidate.display_path,
            "format": candidate.kind,
            "size": candidate.size,
            "read": read_metadata,
            "table": profile,
        }
        _plot_table_primary(
            frame,
            label=label,
            stem=candidate.path.stem,
            source=candidate.display_path,
            plots=plots,
            limits=limits,
        )
        return result, (frame, label, candidate.path.stem)

    sheets, read_metadata = _read_excel(candidate.path, limits)
    sheet_profiles: list[dict[str, Any]] = []
    metadata_by_name = {
        item["name"]: item for item in read_metadata["sheet_metadata"]
    }
    for sheet_name, frame in sheets:
        metadata = metadata_by_name.get(sheet_name, {})
        rows_total = metadata.get("declared_rows")
        if isinstance(rows_total, int) and rows_total:
            rows_total = max(0, rows_total - 1)
        sheet_profiles.append(
            {
                "name": sheet_name,
                "table": _profile_frame(
                    frame,
                    rows_total=rows_total,
                    truncated=bool(metadata.get("truncated")),
                ),
            }
        )
    first_sheet, first_frame = sheets[0]
    sheet_label = f"{label} / {first_sheet}"
    _plot_table_primary(
        first_frame,
        label=sheet_label,
        stem=candidate.path.stem,
        source=candidate.display_path,
        plots=plots,
        limits=limits,
    )
    return (
        {
            "path": candidate.display_path,
            "format": "excel",
            "size": candidate.size,
            "read": read_metadata,
            "sheets": sheet_profiles,
        },
        (first_frame, sheet_label, candidate.path.stem),
    )


def _safe_markdown_text(value: Any) -> str:
    text = str(value)
    text = "".join(character if character.isprintable() else "�" for character in text)
    return text.replace("`", "ˋ")


def _inline_code(value: Any) -> str:
    return f"`{_safe_markdown_text(value)}`"


def _format_size(size: Any) -> str:
    value = max(0, int(size or 0))
    units = ("B", "KB", "MB", "GB")
    amount = float(value)
    unit = units[0]
    for candidate in units:
        unit = candidate
        if amount < 1024 or candidate == units[-1]:
            break
        amount /= 1024
    return f"{int(amount)} {unit}" if unit == "B" else f"{amount:.1f} {unit}"


def _tree_lines(root: dict[str, Any], entries: list[dict[str, Any]]) -> list[str]:
    """Render safe logical paths as a compact tree for the Markdown summary."""
    tree: dict[str, Any] = {"children": {}, "entry": root}
    for entry in entries:
        normalized = _safe_relative_manifest_path(entry.get("path"))
        if normalized is None or normalized == ".":
            continue
        node = tree
        for part in PurePosixPath(normalized).parts:
            node = node["children"].setdefault(
                part, {"children": {}, "entry": None}
            )
        node["entry"] = entry

    def label(name: str, node: dict[str, Any]) -> str:
        entry = node.get("entry") or {}
        entry_type = entry.get("type")
        rendered = _safe_markdown_text(name)
        if entry_type == "directory" or (node["children"] and not entry_type):
            return f"{rendered}/"
        if entry_type == "archive":
            file_format = str(entry.get("format", "archive")).upper()
            return f"{rendered}  [{file_format} 压缩包，{_format_size(entry.get('size'))}]"
        details: list[str] = []
        if entry.get("format"):
            details.append(str(entry["format"]).upper())
        if "size" in entry:
            details.append(_format_size(entry["size"]))
        return f"{rendered}  [{', '.join(details)}]" if details else rendered

    root_name = _safe_markdown_text(root.get("name", "dataset"))
    root_type = root.get("type")
    if root_type == "directory":
        root_label = f"{root_name}/"
    elif root_type == "archive":
        root_label = (
            f"{root_name}  [{str(root.get('format', 'archive')).upper()} 压缩包，"
            f"{_format_size(root.get('size'))}]"
        )
    else:
        details = [str(root.get("format", "file")).upper()]
        if "size" in root:
            details.append(_format_size(root["size"]))
        root_label = f"{root_name}  [{', '.join(details)}]"
    lines = [root_label]

    def render_children(node: dict[str, Any], prefix: str) -> None:
        children = sorted(
            node["children"].items(),
            key=lambda item: (
                0
                if item[1]["children"]
                or (item[1].get("entry") or {}).get("type") == "directory"
                else 1,
                item[0].casefold(),
            ),
        )
        for index, (name, child) in enumerate(children):
            last = index == len(children) - 1
            lines.append(f"{prefix}{'└── ' if last else '├── '}{label(name, child)}")
            if child["children"]:
                render_children(child, f"{prefix}{'    ' if last else '│   '}")

    render_children(tree, "")
    return lines


def _compact_column_evidence(column: dict[str, Any]) -> dict[str, Any]:
    evidence = {
        key: column.get(key)
        for key in (
            "name",
            "dtype",
            "non_null",
            "missing",
            "missing_percent",
            "unique_sampled",
        )
    }
    if isinstance(column.get("statistics"), dict):
        evidence["statistics"] = column["statistics"]
    if isinstance(column.get("range"), dict):
        evidence["range"] = column["range"]
    if isinstance(column.get("top_values"), list):
        evidence["top_values"] = [
            {
                "value": str(item.get("value") or "")[:80],
                "count": item.get("count"),
            }
            for item in column["top_values"][:1]
            if isinstance(item, dict)
        ]
    return evidence


def _compact_table_evidence(table: dict[str, Any]) -> dict[str, Any]:
    columns = table.get("columns")
    columns = columns if isinstance(columns, list) else []
    return {
        "rows_sampled": table.get("rows_sampled"),
        "rows_total": table.get("rows_total"),
        "columns_profiled": table.get("columns_profiled"),
        "truncated": bool(table.get("truncated")),
        "columns": [
            _compact_column_evidence(column)
            for column in columns[:TOOL_EVIDENCE_MAX_COLUMNS]
            if isinstance(column, dict)
        ],
        "columns_omitted": max(0, len(columns) - TOOL_EVIDENCE_MAX_COLUMNS),
    }


def _explicit_table_dimensions(
    path: str,
    table: dict[str, Any],
    *,
    sheet: str | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    columns = table.get("columns")
    columns = columns if isinstance(columns, list) else []
    temporal: list[dict[str, Any]] = []
    latitude: list[str] = []
    longitude: list[str] = []
    for column in columns:
        if not isinstance(column, dict):
            continue
        name = str(column.get("name") or "")
        dtype = str(column.get("dtype") or "")
        if dtype == "datetime" or (
            _TEMPORAL_FIELD_RE.search(name)
            and int(column.get("unique_sampled") or 0) > 1
        ):
            temporal.append(
                {
                    "path": path,
                    "sheet": sheet,
                    "field": name,
                    "dtype": dtype,
                    "range": column.get("range"),
                }
            )
        if _LATITUDE_FIELD_RE.search(f" {name} "):
            latitude.append(name)
        if _LONGITUDE_FIELD_RE.search(f" {name} "):
            longitude.append(name)

    spatial: list[dict[str, Any]] = []
    if latitude and longitude:
        spatial.append(
            {
                "path": path,
                "sheet": sheet,
                "type": "coordinate_fields",
                "latitude_fields": latitude[:2],
                "longitude_fields": longitude[:2],
            }
        )
    return temporal, spatial


def _quicklook_evidence(manifest: dict[str, Any]) -> dict[str, Any]:
    """Return compact, path-safe evidence for the next model decision.

    The persisted manifest remains complete. This projection is deliberately
    small enough for one tool result so a specific multi-part question can be
    answered without a redundant file-read round trip.
    """
    raw_datasets = manifest.get("datasets")
    raw_datasets = raw_datasets if isinstance(raw_datasets, list) else []
    datasets: list[dict[str, Any]] = []
    temporal_dimensions: list[dict[str, Any]] = []
    spatial_dimensions: list[dict[str, Any]] = []

    for dataset in raw_datasets[:TOOL_EVIDENCE_MAX_DATASETS]:
        if not isinstance(dataset, dict):
            continue
        path = str(dataset.get("path") or "")[:512]
        kind = str(dataset.get("format") or "")
        compact: dict[str, Any] = {
            "path": path,
            "format": kind,
            "size": dataset.get("size"),
        }
        if kind == "geotiff":
            compact_bands = []
            for band in dataset.get("bands") or []:
                if not isinstance(band, dict):
                    continue
                compact_bands.append(
                    {
                        "band": band.get("band"),
                        "description": str(band.get("description") or "")[:160] or None,
                        "declared_nodata": band.get("declared_nodata"),
                        "declared_unit": band.get("declared_unit"),
                        "mask_provenance": band.get("mask_provenance"),
                        "masked_count": band.get("masked_count"),
                        "nan_count": band.get("nan_count"),
                        "zero_count": band.get("zero_count"),
                        "valid_zero_count": band.get("valid_zero_count"),
                        "valid_pixels_sampled": band.get("valid_pixels_sampled"),
                        "min": band.get("min"),
                        "max": band.get("max"),
                        "mean": band.get("mean"),
                        "std": band.get("std"),
                    }
                )
            compact.update(
                {
                    "width": dataset.get("width"),
                    "height": dataset.get("height"),
                    "band_count": dataset.get("band_count"),
                    "bands_profiled": dataset.get("bands_profiled"),
                    "dtypes": (dataset.get("dtypes") or [])[:16],
                    "crs": dataset.get("crs"),
                    "bounds": dataset.get("bounds"),
                    "nodata": dataset.get("nodata"),
                    "declared_nodata": dataset.get("declared_nodata"),
                    "declared_unit": dataset.get("declared_unit"),
                    "mask_provenance": dataset.get("mask_provenance"),
                    "masked_count": dataset.get("masked_count"),
                    "nan_count": dataset.get("nan_count"),
                    "zero_count": dataset.get("zero_count"),
                    "valid_zero_count": dataset.get("valid_zero_count"),
                    "sampling": dataset.get("sampling"),
                    "bands": compact_bands,
                    "spatial_profile": dataset.get("spatial_profile"),
                }
            )
            spatial_dimensions.append(
                {
                    "path": path,
                    "type": "raster_grid",
                    "georeferenced": bool(dataset.get("crs")),
                    "crs": dataset.get("crs"),
                    "bounds": dataset.get("bounds"),
                }
            )
            # Raster time is explicit only when a band description identifies a
            # temporal coordinate. A date-looking filename is not enough.
            for band in dataset.get("bands") or []:
                if not isinstance(band, dict):
                    continue
                description = str(band.get("description") or "")[:160]
                if description and _TEMPORAL_FIELD_RE.search(description):
                    temporal_dimensions.append(
                        {
                            "path": path,
                            "band": band.get("band"),
                            "description": description,
                        }
                    )
        elif kind == "excel":
            sheets = dataset.get("sheets")
            sheets = sheets if isinstance(sheets, list) else []
            compact_sheets: list[dict[str, Any]] = []
            for sheet in sheets[:TOOL_EVIDENCE_MAX_SHEETS]:
                if not isinstance(sheet, dict) or not isinstance(sheet.get("table"), dict):
                    continue
                sheet_name = str(sheet.get("name") or "")
                table = sheet["table"]
                compact_sheets.append(
                    {"name": sheet_name, "table": _compact_table_evidence(table)}
                )
                temporal, spatial = _explicit_table_dimensions(
                    path,
                    table,
                    sheet=sheet_name,
                )
                temporal_dimensions.extend(temporal)
                spatial_dimensions.extend(spatial)
            for sheet in sheets[TOOL_EVIDENCE_MAX_SHEETS:]:
                if not isinstance(sheet, dict) or not isinstance(sheet.get("table"), dict):
                    continue
                temporal, spatial = _explicit_table_dimensions(
                    path,
                    sheet["table"],
                    sheet=str(sheet.get("name") or ""),
                )
                temporal_dimensions.extend(temporal)
                spatial_dimensions.extend(spatial)
            compact["sheets"] = compact_sheets
            compact["sheets_omitted"] = max(
                0,
                len(sheets) - TOOL_EVIDENCE_MAX_SHEETS,
            )
        elif isinstance(dataset.get("table"), dict):
            table = dataset["table"]
            compact["table"] = _compact_table_evidence(table)
            temporal, spatial = _explicit_table_dimensions(path, table)
            temporal_dimensions.extend(temporal)
            spatial_dimensions.extend(spatial)
        datasets.append(compact)

    # Capability coverage must reflect every analyzed dataset, even though the
    # value projection above is intentionally limited for model-context size.
    for dataset in raw_datasets[TOOL_EVIDENCE_MAX_DATASETS:]:
        if not isinstance(dataset, dict):
            continue
        path = str(dataset.get("path") or "")[:512]
        kind = str(dataset.get("format") or "")
        if kind == "geotiff":
            spatial_dimensions.append(
                {
                    "path": path,
                    "type": "raster_grid",
                    "georeferenced": bool(dataset.get("crs")),
                    "crs": dataset.get("crs"),
                    "bounds": dataset.get("bounds"),
                }
            )
            for band in dataset.get("bands") or []:
                if not isinstance(band, dict):
                    continue
                description = str(band.get("description") or "")[:160]
                if description and _TEMPORAL_FIELD_RE.search(description):
                    temporal_dimensions.append(
                        {
                            "path": path,
                            "band": band.get("band"),
                            "description": description,
                        }
                    )
        elif kind == "excel":
            for sheet in dataset.get("sheets") or []:
                if not isinstance(sheet, dict) or not isinstance(sheet.get("table"), dict):
                    continue
                temporal, spatial = _explicit_table_dimensions(
                    path,
                    sheet["table"],
                    sheet=str(sheet.get("name") or ""),
                )
                temporal_dimensions.extend(temporal)
                spatial_dimensions.extend(spatial)
        elif isinstance(dataset.get("table"), dict):
            temporal, spatial = _explicit_table_dimensions(path, dataset["table"])
            temporal_dimensions.extend(temporal)
            spatial_dimensions.extend(spatial)

    errors = manifest.get("errors")
    errors = errors if isinstance(errors, list) else []
    return {
        "summary": manifest.get("summary"),
        "limits": manifest.get("limits"),
        "discovery": manifest.get("discovery"),
        "datasets": datasets,
        "datasets_omitted": max(0, len(raw_datasets) - TOOL_EVIDENCE_MAX_DATASETS),
        "errors": [
            {
                "path": str(item.get("path") or "")[:512],
                "error": str(item.get("error") or "")[:512],
            }
            for item in errors[:8]
            if isinstance(item, dict)
        ],
        "errors_omitted": max(0, len(errors) - 8),
        "capabilities": {
            "explicit_temporal_dimensions": temporal_dimensions[:12],
            "explicit_spatial_dimensions": spatial_dimensions[:12],
            "quality_profile_available": bool(datasets),
            "coverage_rule": (
                "Only explicit fields, sheets, coordinates, raster bands, CRS, and sampled values "
                "support analysis. Filenames and catalog period labels do not create a temporal axis."
            ),
        },
    }


def _format_statistic(value: Any) -> str:
    if value is None:
        return "未知"
    if isinstance(value, float):
        return f"{value:.6g}"
    return str(value)


def _table_markdown_evidence(table: dict[str, Any], *, maximum_columns: int = 12) -> list[str]:
    rows_sampled = table.get("rows_sampled")
    rows_total = table.get("rows_total")
    coverage = f"抽样 {rows_sampled} 行"
    if isinstance(rows_total, int):
        coverage += f" / 总计 {rows_total} 行"
    if table.get("truncated"):
        coverage += "（有界抽样，未覆盖全部记录或字段）"
    lines = [f"- 覆盖范围：{coverage}。"]
    columns = table.get("columns")
    columns = columns if isinstance(columns, list) else []
    for column in columns[:maximum_columns]:
        if not isinstance(column, dict):
            continue
        name = _inline_code(column.get("name") or "未命名字段")
        missing = _format_statistic(column.get("missing_percent"))
        detail = f"{name}：类型 {column.get('dtype') or '未知'}，缺失率 {missing}%"
        statistics = column.get("statistics")
        if isinstance(statistics, dict):
            detail += (
                "，最小/均值/中位数/最大值 "
                f"{_format_statistic(statistics.get('min'))} / "
                f"{_format_statistic(statistics.get('mean'))} / "
                f"{_format_statistic(statistics.get('median'))} / "
                f"{_format_statistic(statistics.get('max'))}"
            )
        elif isinstance(column.get("range"), dict):
            detail += (
                f"，范围 {_format_statistic(column['range'].get('min'))} 至 "
                f"{_format_statistic(column['range'].get('max'))}"
            )
        lines.append(f"- {detail}。")
    if len(columns) > maximum_columns:
        lines.append(f"- 其余 {len(columns) - maximum_columns} 个字段详见 JSON 清单。")
    return lines


def _write_markdown(manifest: dict[str, Any], target: Path) -> None:
    summary = manifest["summary"]
    organization = manifest["file_organization"]
    original_tree = organization["original_tree"]
    lines = [
        "# 数据集快速探查",
        "",
        f"- 已分析文件：{summary['files_analyzed']} 个",
        f"- 已生成图表：{summary['plot_count']} 张",
        f"- 抽样策略：表格最多 {manifest['limits']['max_rows_per_table']} 行、"
        f"{manifest['limits']['max_columns']} 列；栅格每波段最多 "
        f"{manifest['limits']['max_raster_pixels']} 个像元。",
        "",
        "## 文件组织结构",
        "",
        "### 原始目录",
        "",
        "```text",
        *_tree_lines(organization["root"], original_tree["entries"]),
        "```",
    ]
    if original_tree["truncated"]:
        lines.extend(
            [
                "",
                f"> 原始目录内容较多，仅展示前 {original_tree['entries_listed']} 项。",
            ]
        )

    lines.extend(["", "### 压缩包层级", ""])
    if organization["archive_layers"]:
        for layer in organization["archive_layers"]:
            lines.append(
                f"- 第 {layer['level']} 层：{_inline_code(layer['path'])} "
                f"（{str(layer['format']).upper()}）→ 解压到 "
                f"{_inline_code(layer['extracted_to'])}"
            )
        if organization["archive_layers_truncated"]:
            lines.append("- 压缩包数量超过展示上限，其余层级已省略。")
    else:
        lines.append("- 未发现需要解压的 ZIP、RAR 或 7z 压缩包。")

    extracted_tree = organization["extracted_tree"]
    lines.extend(["", "### 解压后文件树", ""])
    if extracted_tree["entries"]:
        lines.extend(
            [
                "```text",
                *_tree_lines(
                    {"name": "解压结果", "type": "directory"},
                    extracted_tree["entries"],
                ),
                "```",
            ]
        )
        if extracted_tree["truncated"]:
            lines.extend(
                [
                    "",
                    f"> 解压后文件较多，仅展示前 {extracted_tree['entries_listed']} 项。",
                ]
            )
    else:
        lines.append("- 当前数据集没有压缩包解压结果。")

    lines.extend(
        [
            "",
            "## 文件概览",
            "",
            "| 文件 | 类型 | 关键信息 |",
            "|---|---|---|",
        ]
    )
    for dataset in manifest["datasets"]:
        path = str(dataset["path"]).replace("|", "\\|")
        if dataset["format"] == "geotiff":
            detail = f"{dataset['width']} × {dataset['height']}，{dataset['band_count']} 波段"
        elif dataset["format"] == "netcdf":
            detail = (
                f"{len(dataset.get('variables') or [])} 个变量，"
                f"抽样变量 {dataset.get('sampled_variable') or '未识别'}"
            )
        elif dataset["format"] == "excel":
            detail = f"{len(dataset['sheets'])} 个工作表已剖析"
        else:
            table = dataset["table"]
            detail = f"抽样 {table['rows_sampled']} 行，{table['columns_profiled']} 列"
        lines.append(f"| {path} | {dataset['format']} | {detail} |")

    lines.extend(["", "## 可核验数据证据", ""])
    evidence_datasets = manifest["datasets"][:TOOL_EVIDENCE_MAX_DATASETS]
    for dataset in evidence_datasets:
        lines.extend([f"### {_inline_code(dataset['path'])}", ""])
        if dataset["format"] == "geotiff":
            sampling = dataset.get("sampling") or {}
            lines.append(
                f"- 栅格结构：{dataset.get('width')} × {dataset.get('height')} 像元，"
                f"{dataset.get('band_count')} 个波段；CRS 为 "
                f"{_inline_code(dataset.get('crs') or '未声明')}，边界为 "
                f"{_inline_code(dataset.get('bounds'))}。"
            )
            lines.append(
                f"- 质量与抽样：声明的 NoData 为 "
                f"{_inline_code(dataset.get('declared_nodata'))}；声明的单位为 "
                f"{_inline_code(dataset.get('declared_unit') or 'null（未声明）')}；"
                f"掩膜来源为 "
                f"{_inline_code(dataset.get('mask_provenance'))}；"
                f"每波段抽样 {sampling.get('pixels_per_band')} 个像元"
                f"{'（有界抽样）' if sampling.get('truncated') else '（全量像元）'}。"
            )
            lines.append(
                f"- 抽样质量计数（已剖析波段合计）：权威掩膜排除 "
                f"{dataset.get('masked_count')} 个，原始 NaN {dataset.get('nan_count')} 个，"
                f"原始零值 {dataset.get('zero_count')} 个，其中未被掩膜、保留为有效值的零值 "
                f"{dataset.get('valid_zero_count')} 个。"
            )
            for band in (dataset.get("bands") or [])[:8]:
                lines.append(
                    f"- 波段 {band.get('band')}（声明单位 "
                    f"{_inline_code(band.get('declared_unit') or 'null（未声明）')}）："
                    "有效抽样像元 "
                    f"{band.get('valid_pixels_sampled')}，最小/均值/最大/标准差为 "
                    f"{_format_statistic(band.get('min'))} / "
                    f"{_format_statistic(band.get('mean'))} / "
                    f"{_format_statistic(band.get('max'))} / "
                    f"{_format_statistic(band.get('std'))}。"
                )
            spatial_profile = dataset.get("spatial_profile") or {}
            quantiles = spatial_profile.get("quantiles") or {}
            zones = spatial_profile.get("zone_means") or {}
            minimum = spatial_profile.get("minimum_location") or {}
            maximum = spatial_profile.get("maximum_location") or {}
            lines.append(
                f"- 空间样本有效覆盖率："
                f"{_format_statistic(spatial_profile.get('valid_fraction_percent'))}%；"
                f"P05/P50/P95 为 {_format_statistic(quantiles.get('p05'))} / "
                f"{_format_statistic(quantiles.get('p50'))} / "
                f"{_format_statistic(quantiles.get('p95'))}。"
            )
            lines.append(
                "- 栅格分区均值（左上/右上/左下/右下/中心）为 "
                f"{_format_statistic(zones.get('upper_left'))} / "
                f"{_format_statistic(zones.get('upper_right'))} / "
                f"{_format_statistic(zones.get('lower_left'))} / "
                f"{_format_statistic(zones.get('lower_right'))} / "
                f"{_format_statistic(zones.get('center'))}；这些方向是像元网格方向，"
                "未自动等同于东南西北。"
            )
            lines.append(
                f"- 抽样最小值位置：行 {minimum.get('row')}、列 {minimum.get('column')}、"
                f"坐标 ({_format_statistic(minimum.get('x'))}, "
                f"{_format_statistic(minimum.get('y'))})；抽样最大值位置：行 "
                f"{maximum.get('row')}、列 {maximum.get('column')}、坐标 "
                f"({_format_statistic(maximum.get('x'))}, "
                f"{_format_statistic(maximum.get('y'))})。"
            )
        elif dataset["format"] == "netcdf":
            lines.append(
                f"- NetCDF 维度：{_inline_code(dataset.get('dimensions') or {})}。"
            )
            lines.append(
                f"- 抽样变量：{_inline_code(dataset.get('sampled_variable') or '未识别')}，"
                f"有效抽样值 {dataset.get('sampled_values', 0)} 个。"
            )
            for variable in (dataset.get("variables") or [])[:TOOL_EVIDENCE_MAX_COLUMNS]:
                lines.append(
                    f"- 变量 {_inline_code(variable.get('name') or '')}："
                    f"维度 {variable.get('dimensions') or []}，形状 {variable.get('shape') or []}，"
                    f"单位 {_inline_code(variable.get('units') or '未声明')}。"
                )
        elif dataset["format"] == "excel":
            for sheet in (dataset.get("sheets") or [])[:TOOL_EVIDENCE_MAX_SHEETS]:
                lines.append(f"- 工作表：{_inline_code(sheet.get('name') or '未命名')}。")
                lines.extend(_table_markdown_evidence(sheet.get("table") or {}))
        else:
            lines.extend(_table_markdown_evidence(dataset.get("table") or {}))
        lines.append("")
    if len(manifest["datasets"]) > TOOL_EVIDENCE_MAX_DATASETS:
        lines.append(
            f"> 证据摘要仅展示前 {TOOL_EVIDENCE_MAX_DATASETS} 个已剖析文件；"
            "其余文件的结构化证据保存在 JSON 清单中。"
        )

    compact_evidence = _quicklook_evidence(manifest)
    capabilities = compact_evidence["capabilities"]
    lines.extend(
        [
            "",
            "## 方法与适用边界",
            "",
            "- 本结果由确定性快速探查生成；表格按行数、字段数和读取字节数有界抽样，"
            "栅格按像元数和波段数有界抽样。具体上限与是否截断见 JSON 清单。",
            "- 数值统计沿用源数据中的数值；源数据没有提供单位时，本报告不会猜测单位。"
            "零值不会被自动视为缺失；只有源 NoData/掩膜声明才会从主统计排除对应像元。"
            "缺失率、NoData 声明、掩膜来源、有效抽样数和读取失败用于描述可见的数据质量，"
            "不等同于完整业务质量审计。",
            "- 图表与相关性只描述样本中的模式，不能单独证明因果关系；需要全量或推断统计时，"
            "应针对明确变量和假设进一步分析。",
        ]
    )
    if not capabilities["explicit_temporal_dimensions"]:
        lines.append(
            "- 当前剖析未识别到显式时间字段或带时间描述的波段，因此不能仅依据文件名、"
            "目录名或资料中的时期标签计算年际、月际或其他时间趋势。"
        )
    else:
        temporal_labels = [
            item.get("field") or item.get("description")
            for item in capabilities["explicit_temporal_dimensions"][:5]
        ]
        lines.append(
            "- 已识别的显式时间维度："
            + "、".join(_inline_code(value) for value in temporal_labels if value)
            + "；趋势分析仍需确认时间粒度、缺测和可比性。"
        )
    if capabilities["explicit_spatial_dimensions"]:
        lines.append(
            "- 已识别栅格网格或坐标字段，可描述空间分布；若 CRS 未声明，位置解释仅限"
            "像元/原始坐标范围，不能安全映射到真实地理位置。"
        )
    if manifest["errors"]:
        lines.extend(["", "## 未分析文件", ""])
        for error in manifest["errors"]:
            lines.append(
                f"- {_inline_code(error['path'])}：{_safe_markdown_text(error['error'])}"
            )
    lines.extend(["", "## 输出图表", ""])
    for artifact in manifest["artifacts"]:
        if artifact["media_type"] == "image/png":
            lines.append(f"- [{artifact['title']}]({artifact['path']})")
    target.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate_quicklook(source: Path, output: Path, limits: Limits | None = None) -> dict[str, Any]:
    """Generate a quicklook directory and return its persisted manifest."""
    limits = limits or Limits()
    source = source.expanduser().resolve(strict=True)
    output = output.expanduser().resolve(strict=False)
    if not source.is_file() and not source.is_dir():
        raise QuicklookError("input must be a regular file or directory")
    if source.is_file() and source.stat().st_size > limits.max_source_bytes:
        raise QuicklookError(
            f"source exceeds size limit ({source.stat().st_size} > "
            f"{limits.max_source_bytes} bytes)"
        )
    if output.exists():
        raise QuicklookError(f"output already exists: {output}")
    original_tree = _original_source_tree(source, limits)
    output.parent.mkdir(parents=True, exist_ok=True)
    staging = Path(tempfile.mkdtemp(prefix=f".{output.name}.tmp-", dir=output.parent))
    deadline = Deadline(limits.max_runtime_seconds)
    deadline.arm()
    archive_manifest: dict[str, Any] | None = None
    archive_organization: dict[str, Any] = {
        "archive_layers": [],
        "archive_layers_truncated": False,
        "extracted_entries": [],
        "extracted_tree_truncated": False,
    }
    pre_analysis_errors: list[dict[str, str]] = []
    temporary_workspaces: list[Path] = []
    try:
        if source.is_file() and archive_kind(source) is not None:
            unpacked = staging / "unpacked"
            temporary_workspaces.append(unpacked)
            try:
                archive_manifest = unpack_recursive(
                    source,
                    unpacked,
                    UnpackLimits(
                        max_depth=5,
                        max_archives=limits.max_archives,
                        max_files=limits.max_files_scan,
                        max_total_bytes=limits.max_archive_expanded_bytes,
                        max_single_file_bytes=min(
                            limits.max_source_bytes, limits.max_archive_expanded_bytes
                        ),
                        timeout_seconds=min(limits.max_runtime_seconds, 600),
                    ),
                )
            except Exception as exc:
                raise QuicklookError(
                    _public_error_text(
                        exc,
                        (
                            (source, source.name),
                            (unpacked, "[内部解压目录]"),
                            (staging, "[内部工作目录]"),
                        ),
                        include_type=False,
                    )
                ) from exc
            deadline.check()
            discovery_state = DiscoveryState(maximum=limits.max_files_scan)
            discovered, _nested_archives = _discover(
                unpacked,
                excluded=staging,
                limits=limits,
                state=discovery_state,
                include_archives=False,
                display_prefix=source.name,
            )
            candidates = discovered[: limits.max_files_analyze]
            discovery = discovery_state.manifest(len(candidates))
            archive_count = int(archive_manifest["summary"]["archive_count"])
            final_file_count = int(archive_manifest["summary"]["file_count"])
            archive_organization = _archive_organization(
                archive_manifest,
                outer_display_path=source.name,
                max_layers=limits.max_tree_entries,
                max_files=limits.max_tree_entries,
            )
            source_description = {
                "name": source.name,
                "type": "archive",
                "archive": {
                    "archive_count": archive_count,
                    "final_file_count": final_file_count,
                    "extracted_file_count": final_file_count
                    + max(0, archive_count - 1),
                    "expanded_bytes": int(
                        archive_manifest["summary"]["expanded_bytes"]
                    ),
                },
            }
        elif source.is_dir():
            discovery_state = DiscoveryState(maximum=limits.max_files_scan)
            direct_candidates, directory_archives = _discover(
                source,
                excluded=staging,
                limits=limits,
                state=discovery_state,
                include_archives=True,
            )
            candidates = direct_candidates[: limits.max_files_analyze]
            archive_candidates: list[Candidate] = []
            directory_archive_report: dict[str, Any] = {
                "discovered": len(directory_archives),
                "processed": 0,
                "failed": 0,
                "skipped": len(directory_archives),
                "archive_count": 0,
                "extracted_file_count": 0,
                "expanded_bytes": 0,
                "items": [],
            }
            if directory_archives:
                archive_workspace = staging / "unpacked_archives"
                temporary_workspaces.append(archive_workspace)
                (
                    archive_candidates,
                    directory_archive_report,
                    pre_analysis_errors,
                    archive_organization,
                ) = _unpack_directory_archives(
                    directory_archives,
                    staging=staging,
                    limits=limits,
                    deadline=deadline,
                    discovery_state=discovery_state,
                    selected_count=len(candidates),
                )
                candidates.extend(archive_candidates)
            discovery = discovery_state.manifest(len(candidates))
            source_description = {
                "name": source.name,
                "type": "directory",
                "directory_archives": directory_archive_report,
            }
        else:
            candidates = [_single_candidate(source, limits)]
            discovery = {
                "files_scanned": 1,
                "supported_files_found": 1,
                "files_selected": 1,
                "truncated": False,
            }
            source_description = {"name": source.name, "type": "file"}
        if not candidates:
            raise QuicklookError("no supported CSV/TSV, Excel, NetCDF, or GeoTIFF files found")

        plots = PlotCollector(staging, max(1, min(limits.max_plots, 4)))
        datasets: list[dict[str, Any]] = []
        errors: list[dict[str, str]] = list(pre_analysis_errors)
        first_table: tuple[pd.DataFrame, str, str, str] | None = None
        for candidate in candidates:
            deadline.check()
            if candidate.size > limits.max_source_bytes:
                errors.append(
                    {
                        "path": candidate.display_path,
                        "error": (
                            f"file exceeds size limit ({candidate.size} > "
                            f"{limits.max_source_bytes} bytes)"
                        ),
                    }
                )
                continue
            try:
                if candidate.kind == "geotiff":
                    datasets.append(_raster_profile_and_plot(candidate, plots, limits))
                elif candidate.kind == "netcdf":
                    datasets.append(_netcdf_profile_and_plot(candidate, plots, limits))
                else:
                    profile, table = _analyze_table_candidate(candidate, plots, limits)
                    datasets.append(profile)
                    if first_table is None and table is not None:
                        frame, label, stem = table
                        first_table = (frame, label, stem, candidate.display_path)
            except QuicklookError as exc:
                errors.append(
                    {
                        "path": candidate.display_path,
                        "error": _public_error_text(
                            exc,
                            (
                                (candidate.path, candidate.display_path),
                                (staging, "[内部工作目录]"),
                                (source, source.name),
                            ),
                            include_type=False,
                        ),
                    }
                )
            except Exception as exc:
                errors.append(
                    {
                        "path": candidate.display_path,
                        "error": _public_error_text(
                            exc,
                            (
                                (candidate.path, candidate.display_path),
                                (staging, "[内部工作目录]"),
                                (source, source.name),
                            ),
                        ),
                    }
                )

        deadline.check()
        if first_table is not None and plots.remaining:
            frame, label, stem, display_path = first_table
            _plot_table_extras(
                frame,
                label=label,
                stem=stem,
                source=display_path,
                plots=plots,
                limits=limits,
            )
        if not datasets:
            details = "; ".join(f"{item['path']}: {item['error']}" for item in errors)
            raise QuicklookError(f"all supported files failed analysis: {details[:2_000]}")
        if not plots.artifacts:
            raise QuicklookError("analysis completed but no visualization could be generated")

        # Extracted source data is only a private working set.  Keeping it in the
        # result would make artifact discovery upload the entire dataset again.
        for workspace in temporary_workspaces:
            shutil.rmtree(workspace, ignore_errors=True)
        deadline.check()

        manifest: dict[str, Any] = {
            "version": 1,
            "success": True,
            "source": source_description,
            "file_organization": {
                "root": original_tree["root"],
                "original_tree": {
                    "entries": original_tree["entries"],
                    "entries_listed": original_tree["entries_listed"],
                    "truncated": original_tree["truncated"],
                },
                "archive_layers": archive_organization["archive_layers"],
                "archive_layers_truncated": archive_organization[
                    "archive_layers_truncated"
                ],
                "extracted_tree": {
                    "entries": archive_organization["extracted_entries"],
                    "entries_listed": len(archive_organization["extracted_entries"]),
                    "truncated": archive_organization[
                        "extracted_tree_truncated"
                    ],
                },
            },
            "limits": asdict(limits),
            "discovery": discovery,
            "summary": {
                "files_analyzed": len(datasets),
                "files_failed": len(errors),
                "plot_count": len(plots.artifacts),
                "elapsed_seconds": round(deadline.elapsed, 3),
            },
            "datasets": datasets,
            "errors": errors,
            "artifacts": list(plots.artifacts),
        }
        manifest["artifacts"].extend(
            [
                {
                    "path": "quicklook_summary.md",
                    "media_type": "text/markdown",
                    "role": "summary",
                    "title": "数据集快速探查摘要",
                },
                {
                    "path": "quicklook_manifest.json",
                    "media_type": "application/json",
                    "role": "manifest",
                    "title": "数据集快速探查清单",
                },
            ]
        )
        _write_markdown(manifest, staging / "quicklook_summary.md")
        (staging / "quicklook_manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
            encoding="utf-8",
        )
        os.replace(staging, output)
        return manifest
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    finally:
        deadline.cancel()


def parse_arguments(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Generate a bounded CSV/TSV, Excel, NetCDF, or GeoTIFF quicklook with 1-4 PNG "
            "charts plus JSON and Markdown summaries. Archives are unpacked safely."
        )
    )
    parser.add_argument("input", type=Path, help="dataset file, archive, or directory")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("/home/ubuntu/output/dataset_quicklook"),
        help="new directory for quicklook artifacts",
    )
    parser.add_argument("--max-files", type=int, default=24)
    parser.add_argument("--max-archives", type=int, default=100)
    parser.add_argument("--max-rows", type=int, default=20_000)
    parser.add_argument("--max-columns", type=int, default=80)
    parser.add_argument("--max-plots", type=int, default=4)
    parser.add_argument("--timeout-seconds", type=int, default=90)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_arguments(argv)
    limits = Limits(
        max_files_analyze=max(1, min(args.max_files, 100)),
        max_archives=max(1, min(args.max_archives, 500)),
        max_rows_per_table=max(10, min(args.max_rows, 100_000)),
        max_columns=max(1, min(args.max_columns, 256)),
        max_plots=max(1, min(args.max_plots, 4)),
        max_runtime_seconds=max(5, min(args.timeout_seconds, 600)),
    )
    try:
        manifest = generate_quicklook(args.input, args.output, limits)
    except (OSError, UnpackError, QuicklookError) as exc:
        public_error = _public_error_text(
            exc,
            (
                (args.input, args.input.name or "dataset"),
                (args.output, "[输出目录]"),
            ),
            include_type=False,
        )
        print(
            json.dumps(
                {"success": False, "error": public_error},
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 2
    except Exception as exc:  # Keep failures machine-readable for tool callers.
        public_error = _public_error_text(
            exc,
            (
                (args.input, args.input.name or "dataset"),
                (args.output, "[输出目录]"),
            ),
        )
        print(
            json.dumps(
                {"success": False, "error": public_error},
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 2
    print(
        json.dumps(
            {
                "success": True,
                "output": str(args.output),
                "summary": manifest["summary"],
                "evidence": _quicklook_evidence(manifest),
                "files": [item["path"] for item in manifest["artifacts"]],
                "artifacts": [
                    {
                        key: item.get(key)
                        for key in ("path", "title", "role", "media_type")
                    }
                    for item in manifest["artifacts"]
                    if isinstance(item, dict)
                ],
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
